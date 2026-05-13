"""
Script: Convertir Dataset Cixtur (Excel) a JSONL para RAG
Lifextreme AI - Knowledge Brain
"""
import openpyxl
import json
import os

EXCEL_PATH = r"C:\Users\ASUS\OneDrive\VARIOS\Documentos\GPTS IA\TOURBOT\ENTRENAMIENTO\DATA SET CIXTUR 11_11_25.xlsx"
OUTPUT_PATH = r"C:\Users\ASUS\OneDrive\VARIOS\Documentos\GPTS IA\BIOVET AI\Lifextreme-Web-AI\data\cixtur_knowledge.jsonl"

# Crear carpeta si no existe
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

print("[INFO] Leyendo Dataset Cixtur...")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)

total_registros = 0
total_hojas = 0

with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        registros_hoja = 0
        encontro_header = False
        
        for row in ws.iter_rows(values_only=True):
            # Detectar fila de encabezados
            if row and len(row) >= 2 and str(row[0]).strip().upper() == "PROMPT":
                encontro_header = True
                continue
            
            # Procesar filas de datos
            if encontro_header and row and row[0] and row[1]:
                prompt = str(row[0]).strip()
                completion = str(row[1]).strip()
                
                # Filtrar filas vacías o inválidas
                if len(prompt) > 5 and len(completion) > 5:
                    registro = {
                        "source": sheet_name.strip(),
                        "prompt": prompt,
                        "completion": completion,
                        "type": "qa",
                        "region": "Cusco",
                        "language": "es"
                    }
                    f.write(json.dumps(registro, ensure_ascii=False) + "\n")
                    registros_hoja += 1
                    total_registros += 1
        
        if registros_hoja > 0:
            print(f"  ✅ {sheet_name}: {registros_hoja} registros")
            total_hojas += 1

print(f"\n🎉 CONVERSIÓN COMPLETA:")
print(f"   📁 Hojas procesadas: {total_hojas}/{len(wb.sheetnames)}")
print(f"   📝 Total registros: {total_registros}")
print(f"   💾 Guardado en: {OUTPUT_PATH}")

# Mostrar tamaño del archivo
size = os.path.getsize(OUTPUT_PATH)
print(f"   📦 Tamaño: {size:,} bytes ({size/1024:.1f} KB)")
print(f"\n✅ Listo para subir a Google Cloud!")
